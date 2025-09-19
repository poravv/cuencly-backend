"""
Exportador completo de facturas - Captura TODOS los datos relevantes
Ideal para análisis detallado, auditorías y clientes que necesitan información completa
"""

from __future__ import annotations

import os
import logging
from typing import List, Optional, Dict, Any
from datetime import datetime

import pandas as pd
from decimal import Decimal, ROUND_HALF_UP
import threading

from app.models.models import InvoiceData, ExcelFileInfo
from app.config.settings import settings
from .utils import (
    group_invoices_by_month,
    monthly_excel_path,
    parse_monto,
    formatear_cdc,
    list_excel_files,
    q0,
)

logger = logging.getLogger(__name__)
_EXPORT_LOCK = threading.Lock()

class ExcelExporterCompleto:
    """
    Exportador completo que captura TODOS los datos de la factura XML/OpenAI
    Genera múltiples hojas con información detallada:
    - "Facturas_Completas": Datos principales + campos adicionales
    - "Productos_Detalle": Detalle completo de productos/servicios  
    - "Empresa_Emisor": Información completa del emisor
    - "Cliente_Receptor": Información completa del cliente
    - "Datos_Tecnicos": CDC, timbrados, metadatos técnicos
    - "Resumen_Mensual": Estadísticas y totales del mes
    """

    def __init__(self, output_dir: Optional[str] = None) -> None:
        self.output_dir = output_dir or settings.EXCEL_OUTPUT_DIR or "/app/data/excels"
        # Subdirectorio para exportaciones completas
        self.complete_dir = os.path.join(self.output_dir, "completo")
        os.makedirs(self.complete_dir, exist_ok=True)
        logger.info("ExcelExporterCompleto apuntando a: %s", self.complete_dir)

    def get_monthly_excel_path(self, year_month: str) -> str:
        """Genera ruta para archivo completo"""
        filename = f"facturas_completas_{year_month}.xlsx"
        return os.path.join(self.complete_dir, filename)

    def export_invoices(self, invoices: List[InvoiceData]) -> str:
        """
        Exporta facturas con información completa en múltiples hojas
        """
        if not invoices:
            logger.warning("No hay facturas para exportar (completo)")
            return ""

        try:
            with _EXPORT_LOCK:
                by_month = group_invoices_by_month(invoices)
                last_path = ""
                for ym, invs in by_month.items():
                    logger.info("Procesando %d facturas completas para %s", len(invs), ym)
                    path = self.get_monthly_excel_path(ym)
                    if self._export_month_complete(invs, path, ym):
                        last_path = path
                        logger.info("Archivo Excel completo generado: %s", path)
                return last_path
        except Exception as e:
            logger.error("export_invoices completo error: %s", e, exc_info=True)
            return ""

    def _export_month_complete(self, invoices: List[InvoiceData], excel_path: str, year_month: str) -> bool:
        """Exporta un mes completo con todas las hojas detalladas"""
        try:
            # 1. Hoja principal - Facturas Completas
            facturas_data = self._build_facturas_completas(invoices)
            
            # 2. Productos detallados
            productos_data = self._build_productos_detalle(invoices)
            
            # 3. Información de empresas emisoras
            empresas_data = self._build_empresas_emisor(invoices)
            
            # 4. Información de clientes
            clientes_data = self._build_clientes_receptor(invoices)
            
            # 5. Datos técnicos (CDC, timbrados, etc.)
            tecnicos_data = self._build_datos_tecnicos(invoices)
            
            # 6. Resumen estadístico
            resumen_data = self._build_resumen_mensual(invoices, year_month)

            # Convertir a DataFrames
            df_facturas = pd.DataFrame(facturas_data)
            df_productos = pd.DataFrame(productos_data)
            df_empresas = pd.DataFrame(empresas_data)
            df_clientes = pd.DataFrame(clientes_data)
            df_tecnicos = pd.DataFrame(tecnicos_data)
            df_resumen = pd.DataFrame(resumen_data)

            # Merge con datos existentes si el archivo ya existe
            if os.path.exists(excel_path):
                self._merge_with_existing(excel_path, {
                    'Facturas_Completas': df_facturas,
                    'Productos_Detalle': df_productos,
                    'Empresa_Emisor': df_empresas,
                    'Cliente_Receptor': df_clientes,
                    'Datos_Tecnicos': df_tecnicos,
                    'Resumen_Mensual': df_resumen
                })
            else:
                # Escribir archivo nuevo
                with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                    df_facturas.to_excel(writer, sheet_name="Facturas_Completas", index=False)
                    df_productos.to_excel(writer, sheet_name="Productos_Detalle", index=False)
                    df_empresas.to_excel(writer, sheet_name="Empresa_Emisor", index=False)
                    df_clientes.to_excel(writer, sheet_name="Cliente_Receptor", index=False)
                    df_tecnicos.to_excel(writer, sheet_name="Datos_Tecnicos", index=False)
                    df_resumen.to_excel(writer, sheet_name="Resumen_Mensual", index=False)

            # Aplicar formato profesional
            self._apply_complete_formatting(excel_path)
            
            logger.info("✅ Archivo Excel completo generado con %d facturas: %s", len(df_facturas), excel_path)
            return True

        except Exception as e:
            logger.error("export_month_complete error: %s", e, exc_info=True)
            return False

    def _build_facturas_completas(self, invoices: List[InvoiceData]) -> List[Dict[str, Any]]:
        """Construye datos completos de facturas principales"""
        data = []
        for inv in invoices:
            fecha_str = inv.fecha.strftime("%Y-%m-%d") if inv.fecha else ""
            fecha_procesado = inv.procesado_en.strftime("%Y-%m-%d %H:%M:%S") if getattr(inv, "procesado_en", None) else ""
            
            row = {
                # Identificación
                "id_interno": f"{inv.ruc_emisor or ''}_{inv.numero_factura or ''}_{fecha_str}",
                "fecha_factura": fecha_str,
                "numero_factura": inv.numero_factura or "",
                "tipo_documento": getattr(inv, "tipo_documento", "CO"),
                
                # Emisor
                "ruc_emisor": inv.ruc_emisor or "",
                "nombre_emisor": inv.nombre_emisor or "",
                "actividad_economica": getattr(inv, "actividad_economica", ""),
                
                # Cliente
                "ruc_cliente": inv.ruc_cliente or "",
                "nombre_cliente": inv.nombre_cliente or "",
                "email_cliente": inv.email_cliente or "",
                
                # Montos principales
                "moneda": getattr(inv, "moneda", "GS"),
                "tipo_cambio": float(getattr(inv, "tipo_cambio", 1.0) or 1.0),
                "monto_total": float(getattr(inv, "monto_total", 0) or 0),
                
                # Desglose IVA
                "gravado_10": float(getattr(inv, "gravado_10", inv.subtotal_10) or 0),
                "iva_10": float(getattr(inv, "iva_10", 0) or 0),
                "gravado_5": float(getattr(inv, "gravado_5", inv.subtotal_5) or 0),
                "iva_5": float(getattr(inv, "iva_5", 0) or 0),
                "exentas": float(getattr(inv, "subtotal_exentas", 0) or 0),
                "total_iva": float(getattr(inv, "iva", 0) or 0),
                
                # Condiciones
                "condicion_venta": getattr(inv, "condicion_venta", "CONTADO"),
                "condicion_compra": getattr(inv, "condicion_compra", "CONTADO"),
                
                # Descripción y observaciones
                "descripcion_factura": getattr(inv, "descripcion_factura", ""),
                "observacion": getattr(inv, "observacion", ""),
                
                # Metadatos de procesamiento
                "email_origen": getattr(inv, "email_origen", ""),
                "fecha_procesado": fecha_procesado,
                "mes_proceso": getattr(inv, "mes_proceso", ""),
                "pdf_path": getattr(inv, "pdf_path", ""),
                
                # Contadores
                "cantidad_productos": len(getattr(inv, "productos", []) or []),
                "tiene_cdc": bool(getattr(inv, "cdc", "")),
                "tiene_timbrado": bool(getattr(inv, "timbrado", "")),
            }
            data.append(row)
        return data

    def _build_productos_detalle(self, invoices: List[InvoiceData]) -> List[Dict[str, Any]]:
        """Construye detalle completo de productos"""
        data = []
        for inv in invoices:
            factura_id = f"{inv.ruc_emisor or ''}_{inv.numero_factura or ''}"
            fecha_str = inv.fecha.strftime("%Y-%m-%d") if inv.fecha else ""
            
            productos = getattr(inv, "productos", []) or []
            if not productos:
                # Crear entrada vacía para facturas sin productos detallados
                data.append({
                    "factura_id": factura_id,
                    "numero_factura": inv.numero_factura or "",
                    "fecha_factura": fecha_str,
                    "ruc_emisor": inv.ruc_emisor or "",
                    "item_numero": 0,
                    "articulo": "Sin detalle de productos",
                    "cantidad": 0,
                    "precio_unitario": 0,
                    "total_item": 0,
                    "iva_aplicable": 0,
                    "moneda": getattr(inv, "moneda", "GS"),
                })
            else:
                for idx, prod in enumerate(productos, 1):
                    if isinstance(prod, dict):
                        articulo = prod.get("articulo", "")
                        cantidad = float(prod.get("cantidad", 0) or 0)
                        precio_u = float(prod.get("precio_unitario", 0) or 0)
                        total = float(prod.get("total", 0) or 0)
                        iva = int(prod.get("iva", 0) or 0)
                    else:
                        articulo = getattr(prod, "articulo", "") or ""
                        cantidad = float(getattr(prod, "cantidad", 0) or 0)
                        precio_u = float(getattr(prod, "precio_unitario", 0) or 0)
                        total = float(getattr(prod, "total", 0) or 0)
                        iva = int(getattr(prod, "iva", 0) or 0)

                    data.append({
                        "factura_id": factura_id,
                        "numero_factura": inv.numero_factura or "",
                        "fecha_factura": fecha_str,
                        "ruc_emisor": inv.ruc_emisor or "",
                        "item_numero": idx,
                        "articulo": articulo,
                        "cantidad": cantidad,
                        "precio_unitario": precio_u,
                        "total_item": total,
                        "iva_aplicable": iva,
                        "moneda": getattr(inv, "moneda", "GS"),
                    })
        return data

    def _build_empresas_emisor(self, invoices: List[InvoiceData]) -> List[Dict[str, Any]]:
        """Construye información completa de empresas emisoras"""
        empresas_vistas = set()
        data = []
        
        for inv in invoices:
            ruc = inv.ruc_emisor or ""
            if not ruc or ruc in empresas_vistas:
                continue
                
            empresas_vistas.add(ruc)
            empresa_data = getattr(inv, "empresa", None)
            
            row = {
                "ruc": ruc,
                "nombre": inv.nombre_emisor or "",
                "actividad_economica": getattr(inv, "actividad_economica", ""),
                
                # Datos adicionales de empresa si están disponibles
                "direccion": empresa_data.direccion if empresa_data else "",
                "telefono": empresa_data.telefono if empresa_data else "",
                "actividad_detallada": empresa_data.actividad_economica if empresa_data else "",
                
                # Estadísticas
                "total_facturas": sum(1 for i in invoices if (i.ruc_emisor or "") == ruc),
                "monto_total_periodo": sum(float(getattr(i, "monto_total", 0) or 0) 
                                         for i in invoices if (i.ruc_emisor or "") == ruc),
                "primera_factura": min((i.fecha for i in invoices if (i.ruc_emisor or "") == ruc and i.fecha), 
                                     default=None),
                "ultima_factura": max((i.fecha for i in invoices if (i.ruc_emisor or "") == ruc and i.fecha), 
                                    default=None),
            }
            
            # Formatear fechas
            if row["primera_factura"]:
                row["primera_factura"] = row["primera_factura"].strftime("%Y-%m-%d")
            if row["ultima_factura"]:
                row["ultima_factura"] = row["ultima_factura"].strftime("%Y-%m-%d")
                
            data.append(row)
        return data

    def _build_clientes_receptor(self, invoices: List[InvoiceData]) -> List[Dict[str, Any]]:
        """Construye información de clientes/receptores"""
        clientes_vistos = set()
        data = []
        
        for inv in invoices:
            ruc_cliente = inv.ruc_cliente or ""
            if not ruc_cliente or ruc_cliente in clientes_vistos:
                continue
                
            clientes_vistos.add(ruc_cliente)
            cliente_data = getattr(inv, "cliente", None)
            
            row = {
                "ruc_cliente": ruc_cliente,
                "nombre_cliente": inv.nombre_cliente or "",
                "email_cliente": inv.email_cliente or "",
                
                # Datos adicionales del cliente si están disponibles
                "email_adicional": cliente_data.email if cliente_data else "",
                
                # Estadísticas del cliente
                "total_facturas_recibidas": sum(1 for i in invoices if (i.ruc_cliente or "") == ruc_cliente),
                "monto_total_compras": sum(float(getattr(i, "monto_total", 0) or 0) 
                                         for i in invoices if (i.ruc_cliente or "") == ruc_cliente),
                "primera_compra": min((i.fecha for i in invoices if (i.ruc_cliente or "") == ruc_cliente and i.fecha), 
                                    default=None),
                "ultima_compra": max((i.fecha for i in invoices if (i.ruc_cliente or "") == ruc_cliente and i.fecha), 
                                   default=None),
            }
            
            # Formatear fechas
            if row["primera_compra"]:
                row["primera_compra"] = row["primera_compra"].strftime("%Y-%m-%d")
            if row["ultima_compra"]:
                row["ultima_compra"] = row["ultima_compra"].strftime("%Y-%m-%d")
                
            data.append(row)
        return data

    def _build_datos_tecnicos(self, invoices: List[InvoiceData]) -> List[Dict[str, Any]]:
        """Construye datos técnicos: CDC, timbrados, validaciones"""
        data = []
        for inv in invoices:
            timbrado_data = getattr(inv, "timbrado_data", None)
            factura_data = getattr(inv, "factura_data", None)
            
            cdc = getattr(inv, "cdc", "") or (factura_data.cdc if factura_data else "")
            timbrado = getattr(inv, "timbrado", "") or (timbrado_data.nro if timbrado_data else "")
            
            row = {
                "factura_id": f"{inv.ruc_emisor or ''}_{inv.numero_factura or ''}",
                "numero_factura": inv.numero_factura or "",
                "fecha_factura": inv.fecha.strftime("%Y-%m-%d") if inv.fecha else "",
                
                # CDC y validaciones
                "cdc": formatear_cdc(cdc),
                "cdc_valido": len(str(cdc).replace("-", "").replace(" ", "")) == 44 and str(cdc).replace("-", "").replace(" ", "").isdigit(),
                "cdc_formateado": bool(cdc and ("-" in str(cdc) or " " in str(cdc))),
                
                # Timbrado
                "timbrado": timbrado,
                "timbrado_fecha_inicio": timbrado_data.fecha_inicio_vigencia if timbrado_data else "",
                "timbrado_valido_hasta": timbrado_data.valido_hasta if timbrado_data else "",
                "timbrado_vigente": self._validar_timbrado_vigente(timbrado_data, inv.fecha),
                
                # Datos adicionales de factura
                "contado_nro": factura_data.contado_nro if factura_data else "",
                "caja_nro": factura_data.caja_nro if factura_data else "",
                "condicion_factura": factura_data.condicion_venta if factura_data else "",
                
                # Metadatos de procesamiento
                "fuente_procesamiento": "XML_NATIVO" if cdc else "OPENAI_VISION",
                "calidad_datos": "ALTA" if cdc and timbrado else "MEDIA",
                "requiere_revision": not bool(cdc and timbrado),
            }
            data.append(row)
        return data

    def _build_resumen_mensual(self, invoices: List[InvoiceData], year_month: str) -> List[Dict[str, Any]]:
        """Construye resumen estadístico mensual"""
        total_facturas = len(invoices)
        facturas_con_cdc = sum(1 for inv in invoices if getattr(inv, "cdc", ""))
        facturas_con_timbrado = sum(1 for inv in invoices if getattr(inv, "timbrado", ""))
        
        monto_total_gs = sum(float(getattr(inv, "monto_total", 0) or 0) 
                           for inv in invoices if (getattr(inv, "moneda", "GS") or "GS").upper() in ["GS", "PYG"])
        monto_total_usd = sum(float(getattr(inv, "monto_total", 0) or 0) 
                            for inv in invoices if (getattr(inv, "moneda", "GS") or "GS").upper() == "USD")
        
        total_iva = sum(float(getattr(inv, "iva", 0) or 0) for inv in invoices)
        
        proveedores_unicos = len(set(inv.ruc_emisor for inv in invoices if inv.ruc_emisor))
        clientes_unicos = len(set(inv.ruc_cliente for inv in invoices if inv.ruc_cliente))
        
        data = [{
            "periodo": year_month,
            "fecha_generacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            
            # Contadores
            "total_facturas": total_facturas,
            "facturas_con_cdc": facturas_con_cdc,
            "facturas_con_timbrado": facturas_con_timbrado,
            "proveedores_unicos": proveedores_unicos,
            "clientes_unicos": clientes_unicos,
            
            # Porcentajes de calidad
            "porcentaje_cdc": round((facturas_con_cdc / total_facturas * 100) if total_facturas > 0 else 0, 2),
            "porcentaje_timbrado": round((facturas_con_timbrado / total_facturas * 100) if total_facturas > 0 else 0, 2),
            
            # Montos totales
            "monto_total_gs": monto_total_gs,
            "monto_total_usd": monto_total_usd,
            "total_iva": total_iva,
            
            # Promedios
            "promedio_factura_gs": round(monto_total_gs / max(1, total_facturas), 2),
            "promedio_iva": round(total_iva / max(1, total_facturas), 2),
            
            # Distribución por tipo
            "facturas_contado": sum(1 for inv in invoices if (getattr(inv, "condicion_venta", "CONTADO") or "CONTADO").upper() == "CONTADO"),
            "facturas_credito": sum(1 for inv in invoices if (getattr(inv, "condicion_venta", "CONTADO") or "CONTADO").upper() == "CREDITO"),
        }]
        
        return data

    def _validar_timbrado_vigente(self, timbrado_data, fecha_factura) -> bool:
        """Valida si el timbrado está vigente para la fecha de la factura"""
        if not timbrado_data or not fecha_factura:
            return False
            
        try:
            # Aquí podrías agregar lógica de validación de fechas de vigencia
            # Por ahora retorna True si tiene datos
            return bool(timbrado_data.nro and timbrado_data.fecha_inicio_vigencia)
        except Exception:
            return False

    def _merge_with_existing(self, excel_path: str, new_dataframes: Dict[str, pd.DataFrame]):
        """Merge datos nuevos con archivo existente"""
        try:
            # Leer hojas existentes
            existing_dfs = {}
            for sheet_name in new_dataframes.keys():
                try:
                    existing_dfs[sheet_name] = pd.read_excel(excel_path, sheet_name=sheet_name)
                except Exception:
                    existing_dfs[sheet_name] = pd.DataFrame()

            # Combinar datos
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                for sheet_name, new_df in new_dataframes.items():
                    existing_df = existing_dfs.get(sheet_name, pd.DataFrame())
                    
                    if not existing_df.empty:
                        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                        # Deduplicar basado en columnas clave según la hoja
                        if sheet_name == "Facturas_Completas":
                            combined_df.drop_duplicates(subset=["id_interno"], keep="last", inplace=True)
                        elif sheet_name == "Productos_Detalle":
                            combined_df.drop_duplicates(subset=["factura_id", "item_numero"], keep="last", inplace=True)
                        elif sheet_name in ["Empresa_Emisor", "Cliente_Receptor"]:
                            key_col = "ruc" if sheet_name == "Empresa_Emisor" else "ruc_cliente"
                            combined_df.drop_duplicates(subset=[key_col], keep="last", inplace=True)
                        elif sheet_name == "Datos_Tecnicos":
                            combined_df.drop_duplicates(subset=["factura_id"], keep="last", inplace=True)
                    else:
                        combined_df = new_df
                    
                    combined_df.to_excel(writer, sheet_name=sheet_name, index=False)

        except Exception as e:
            logger.error(f"Error merging with existing file: {e}")
            # Fallback: sobrescribir archivo
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                for sheet_name, df in new_dataframes.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

    def _apply_complete_formatting(self, excel_path: str):
        """Aplica formato profesional al archivo Excel completo"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = load_workbook(excel_path)
            
            # Colores corporativos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Formatear headers
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-ajustar columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_path)
            logger.info("Formato aplicado exitosamente al archivo completo")
            
        except Exception as e:
            logger.warning(f"No se pudo aplicar formato: {e}")

    def get_available_excel_files(self) -> List[ExcelFileInfo]:
        """Lista archivos Excel completos disponibles"""
        return list_excel_files(self.complete_dir)

    def get_excel_by_month(self, year_month: str) -> Optional[str]:
        """Obtiene archivo Excel completo por mes"""
        path = self.get_monthly_excel_path(year_month)
        return path if os.path.exists(path) else None