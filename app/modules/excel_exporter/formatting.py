from __future__ import annotations

import logging
from typing import Dict

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger(__name__)

# ---------------------------
# Hoja Resumen
# ---------------------------

def write_summary_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, year_month: str) -> None:
    """
    Crea la hoja "Resumen" (totaliza columnas clave si existen).
    """
    try:
        columns_map = {
            "gra10": "Gravado 10%",
            "iva10": "IVA 10%",
            "gra5": "Gravado 5%",
            "iva5": "IVA 5%",
            "exentos": "Exento",
            "monto_total": "Total Factura",
        }
        totals: Dict[str, float] = {}
        for key, label in columns_map.items():
            totals[label] = float(df[key].sum()) if key in df.columns else 0.0

        resumen_data = [
            ["RESUMEN MENSUAL", ""],
            ["Período", year_month],
            ["", ""],
            ["Total Facturas", len(df)],
            ["", ""],
            ["IMPORTES", ""],
            ["Gravado 10%", totals["Gravado 10%"]],
            ["IVA 10%", totals["IVA 10%"]],
            ["Gravado 5%", totals["Gravado 5%"]],
            ["IVA 5%", totals["IVA 5%"]],
            ["Exento", totals["Exento"]],
            ["", ""],
            ["TOTAL GENERAL", totals["Total Factura"]],
        ]

        resumen_df = pd.DataFrame(resumen_data, columns=["Concepto", "Valor"])
        resumen_df.to_excel(writer, sheet_name="Resumen", index=False)
        logger.info("✅ Hoja de resumen creada para %s", year_month)
    except Exception as e:
        logger.error("Resumen error: %s", e)


# ---------------------------
# Estilos/formatos ASCONT
# ---------------------------

def apply_ascont_formatting(excel_path: str) -> None:
    """
    Aplica estilos a 'Facturas ASCONT' y 'Resumen'.
    """
    try:
        wb = openpyxl.load_workbook(excel_path)

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Facturas ASCONT
        if "Facturas ASCONT" in wb.sheetnames:
            ws = wb["Facturas ASCONT"]

            # encabezados
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # datos
            max_col = ws.max_column
            header_row = [ws.cell(1, c).value for c in range(1, max_col + 1)]
            # Índices de columnas de interés
            try:
                col_idx = {name: header_row.index(name) + 1 for name in header_row}
            except ValueError:
                col_idx = {}

            amount_headers = ["gra10", "iva10", "gra5", "iva5", "exentos", "monto_total"]
            moneda_col = col_idx.get("moneda", None)
            tipo_cambio_col = col_idx.get("tipo_cambio", None)

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # Estética general
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

                # Formato numérico por moneda (GS sin decimales, USD/otras con decimales)
                moneda_val = None
                if moneda_col is not None:
                    moneda_val = ws.cell(row=row[0].row, column=moneda_col).value
                is_gs = str(moneda_val or "GS").upper() in {"GS", "PYG"}

                for hdr in amount_headers:
                    c = col_idx.get(hdr)
                    if not c:
                        continue
                    cell = ws.cell(row=row[0].row, column=c)
                    cell.number_format = "0" if is_gs else "0.00"

                # tipo_cambio: solo tiene sentido con moneda extranjera → 2 decimales; en GS queda 0
                if tipo_cambio_col:
                    tc_cell = ws.cell(row=row[0].row, column=tipo_cambio_col)
                    tc_cell.number_format = "0.00"

            # autosize
            for col in ws.columns:
                values = [len(str(c.value)) for c in col if c.value is not None]
                if not values:
                    continue
                width = min(max(values) + 2, 50)
                ws.column_dimensions[col[0].column_letter].width = width

            ws.freeze_panes = "A2"

        # Productos: aplicar formato numérico por moneda
        if "Productos" in wb.sheetnames:
            wsp = wb["Productos"]
            max_col_p = wsp.max_column
            headers_p = [wsp.cell(1, c).value for c in range(1, max_col_p + 1)]
            col_idx_p = {name: headers_p.index(name) + 1 for name in headers_p if name}

            moneda_col = col_idx_p.get("moneda", None)
            for row in wsp.iter_rows(min_row=2, max_row=wsp.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                is_gs = True
                if moneda_col is not None:
                    mval = wsp.cell(row=row[0].row, column=moneda_col).value
                    is_gs = str(mval or "GS").upper() in {"GS", "PYG"}

                for h in ("cantidad", "precio_unitario", "total"):
                    c = col_idx_p.get(h)
                    if not c:
                        continue
                    cell = wsp.cell(row=row[0].row, column=c)
                    cell.number_format = "0" if is_gs else "0.00"

        # Resumen
        if "Resumen" in wb.sheetnames:
            ws = wb["Resumen"]
            for row in ws.iter_rows():
                for cell in row:
                    text = str(cell.value) if cell.value is not None else ""
                    if cell.row == 1 or "RESUMEN" in text or "TOTAL GENERAL" in text or "IMPORTES" in text:
                        cell.font = Font(bold=True, size=12 if "RESUMEN" in text else 11)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00"

        wb.save(excel_path)
        logger.info("Formato ASCONT aplicado")
    except Exception as e:
        logger.error("apply_ascont_formatting error: %s", e)
